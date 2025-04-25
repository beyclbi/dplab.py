import os
import uuid
import hashlib
from tkinter import simpledialog, messagebox, Tk

LICENSE_FILE = os.path.expanduser("~/.dp_lab_license.key")
LICENSES_TXT = "licenses.txt"  # Geçerli lisans anahtarlarının bulunduğu dosya

def get_device_id():
    return str(uuid.getnode())  # Cihaza özel eşsiz kimlik

def hash_license_key(key):
    return hashlib.sha256(key.encode()).hexdigest()

def is_license_valid(hashed_key):
    if not os.path.exists(LICENSES_TXT):
        return False
    with open(LICENSES_TXT, "r", encoding="utf-8") as f:
        valid_keys = [hash_license_key(line.strip()) for line in f if line.strip()]
    return hashed_key in valid_keys

def save_license(hashed_key):
    with open(LICENSE_FILE, "w") as f:
        f.write(hashed_key)

def load_saved_license():
    if not os.path.exists(LICENSE_FILE):
        return None
    with open(LICENSE_FILE, "r") as f:
        return f.read().strip()

def verify_or_request_license():
    saved_license = load_saved_license()
    if saved_license and is_license_valid(saved_license):
        return True

    root = Tk()
    root.withdraw()  # Arayüzü gizle
    while True:
        license_key = simpledialog.askstring("Lisans Doğrulama", "Lütfen lisans anahtarınızı girin:")
        if license_key is None:
            messagebox.showerror("Lisans Gerekli", "Programı kullanmak için lisans kodu girmelisiniz.")
            exit()

        hashed = hash_license_key(license_key)
        if is_license_valid(hashed):
            save_license(hashed)
            messagebox.showinfo("Lisans Onaylandı", "Lisans doğrulandı. Program başlatılıyor.")
            return True
        else:
            messagebox.showerror("Geçersiz Lisans", "Girilen lisans kodu geçersiz.")


import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from fpdf import FPDF
import openpyxl
import os
import json
from datetime import datetime
import qrcode
from PIL import Image, ImageTk


# === KLASÖRLER ===
os.makedirs("data/qrcodes", exist_ok=True)
FILES = {
    "jobs": "data/jobs.json",
    "clinics": "data/clinics.json",
    "doctors": "data/doctors.json",
    "prices": "data/prices.json"
}

def load_data(file):
    if os.path.exists(file):
        with open(file, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

def save_data(file, data):
    with open(file, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

jobs = load_data(FILES["jobs"])
clinics = load_data(FILES["clinics"])
doctors = load_data(FILES["doctors"])
prices = load_data(FILES["prices"])

if not verify_or_request_license():
    exit()


root = tk.Tk()
root.title("DP Lab - Diş Protez Takip")
root.geometry("1200x750")

notebook = ttk.Notebook(root)
notebook.pack(fill="both", expand=True)

new_job_tab = ttk.Frame(notebook)
all_jobs_tab = ttk.Frame(notebook)
clinic_tab = ttk.Frame(notebook)
price_tab = ttk.Frame(notebook)
finance_tab = ttk.Frame(notebook)
notebook.add(finance_tab, text="Muhasebe")


notebook.add(new_job_tab, text="Yeni İş Girişi")
notebook.add(all_jobs_tab, text="Tüm İşler")
notebook.add(clinic_tab, text="Klinik & Doktor")
notebook.add(price_tab, text="Fiyat Listesi")
report_tab = ttk.Frame(notebook)
notebook.add(report_tab, text="Raporlama")


editing_index = None

# === YENİ İŞ GİRİŞİ ===
name_var = tk.StringVar()
surname_var = tk.StringVar()
clinic_var = tk.StringVar()
job_status_var = tk.StringVar(value="Hazırlanıyor")
doctor_var = tk.StringVar()
prosthesis_var = tk.StringVar()
member_count_var = tk.StringVar(value="1")
note_var = tk.StringVar()
date_var = tk.StringVar(value=datetime.today().strftime("%d/%m/%Y"))
total_price_var = tk.StringVar(value="0.00")

def update_price():
    selected_clinic = clinic_var.get()
    selected_type = prosthesis_var.get()
    found_price = None

    for p in prices:
        if p["type"] == selected_type and p["clinic"] == selected_clinic:
            found_price = p["price"]
            break

    if found_price is None:
        # Klinik için fiyat yoksa 'Genel' fiyatı bul
        for p in prices:
            if p["type"] == selected_type and p["clinic"] == "Genel":
                found_price = p["price"]
                break

    try:
        if found_price:
            total = float(found_price) * int(member_count_var.get())
            total_price_var.set(f"{total:.2f}")
        else:
            total_price_var.set("0.00")
    except:
        total_price_var.set("0.00")


def update_doctor_list(event=None):
    selected_clinic = clinic_var.get()
    relevant_doctors = [d["name"] for d in doctors if d["clinic"] == selected_clinic]
    doctor_combo['values'] = relevant_doctors
    if relevant_doctors:
        doctor_combo.current(0)

def save_job():
    global editing_index
    job = {
        "patient_name": name_var.get(),
        "patient_surname": surname_var.get(),
        "clinic": clinic_var.get(),
        "doctor": doctor_var.get(),
        "prosthesis": prosthesis_var.get(),
        "count": member_count_var.get(),
        "note": note_var.get(),
        "date": date_var.get(),
        "total_price": total_price_var.get(),
        "status": job_status_var.get()
    }
    ...


    
    if editing_index is not None:
        jobs[editing_index] = job
        editing_index = None
    else:
        jobs.append(job)
    save_data(FILES["jobs"], jobs)
    refresh_jobs()
    clear_fields()

def clear_fields():
    name_var.set("")
    surname_var.set("")
    clinic_var.set("")
    doctor_var.set("")
    prosthesis_var.set("")
    member_count_var.set("1")
    note_var.set("")
    date_var.set(datetime.today().strftime("%d/%m/%Y"))
    total_price_var.set("0.00")
    job_status_var.set("Hazırlanıyor")


tk.Label(new_job_tab, text="Ad").grid(row=0, column=0, sticky="w", padx=10, pady=5)
tk.Entry(new_job_tab, textvariable=name_var).grid(row=0, column=1)

tk.Label(new_job_tab, text="Soyad").grid(row=1, column=0, sticky="w", padx=10)
tk.Entry(new_job_tab, textvariable=surname_var).grid(row=1, column=1)

tk.Label(new_job_tab, text="Klinik").grid(row=2, column=0, sticky="w", padx=10)
clinic_combo = ttk.Combobox(new_job_tab, textvariable=clinic_var, values=[c["name"] for c in clinics])
clinic_combo.grid(row=2, column=1)
clinic_combo.bind("<<ComboboxSelected>>", update_doctor_list)

tk.Label(new_job_tab, text="Doktor").grid(row=3, column=0, sticky="w", padx=10)
doctor_combo = ttk.Combobox(new_job_tab, textvariable=doctor_var)
doctor_combo.grid(row=3, column=1)

tk.Label(new_job_tab, text="Protez Tipi").grid(row=4, column=0, sticky="w", padx=10)
prosthesis_combo = ttk.Combobox(new_job_tab, textvariable=prosthesis_var, values=[p["type"] for p in prices])
prosthesis_combo.grid(row=4, column=1)
prosthesis_combo.bind("<<ComboboxSelected>>", lambda e: update_price())

tk.Label(new_job_tab, text="Üye Sayısı").grid(row=5, column=0, sticky="w", padx=10)
tk.Entry(new_job_tab, textvariable=member_count_var).grid(row=5, column=1)

tk.Label(new_job_tab, text="Not").grid(row=6, column=0, sticky="w", padx=10)
tk.Entry(new_job_tab, textvariable=note_var).grid(row=6, column=1)

tk.Label(new_job_tab, text="Tarih").grid(row=7, column=0, sticky="w", padx=10)
tk.Entry(new_job_tab, textvariable=date_var).grid(row=7, column=1)
tk.Label(new_job_tab, text="İş Durumu").grid(row=8, column=0, sticky="w", padx=10)
status_options = ["Hazırlanıyor", "Beklemede", "Yapımda", "Tamamlandı", "Teslim Edildi"]
tk.OptionMenu(new_job_tab, job_status_var, *status_options).grid(row=8, column=1)


tk.Label(new_job_tab, text="Toplam Fiyat (₺)").grid(row=9, column=0, sticky="w", padx=10)
tk.Entry(new_job_tab, textvariable=total_price_var, state="readonly").grid(row=9, column=1)

tk.Button(new_job_tab, text="Kaydet", command=save_job, bg="#007acc", fg="white").grid(row=10, column=1, pady=10)


# === FİLTRE VE TÜM İŞLER ===
filter_clinic = tk.StringVar()
filter_doctor = tk.StringVar()
filter_start = tk.StringVar()
filter_end = tk.StringVar()
filter_name = tk.StringVar()

def filter_jobs():
    result = []
    for job in jobs:
        date_ok = True
        if filter_start.get():
            date_ok &= datetime.strptime(job["date"], "%d/%m/%Y") >= datetime.strptime(filter_start.get(), "%d/%m/%Y")
        if filter_end.get():
            date_ok &= datetime.strptime(job["date"], "%d/%m/%Y") <= datetime.strptime(filter_end.get(), "%d/%m/%Y")
        if filter_clinic.get() and filter_clinic.get() not in job["clinic"]:
            continue
        if filter_doctor.get() and filter_doctor.get() not in job["doctor"]:
            continue
        if filter_name.get():
            full_name = (job["patient_name"] + " " + job["patient_surname"]).lower()
            if filter_name.get().lower() not in full_name:
                continue
        if date_ok:
            result.append(job)
    return result

def refresh_jobs():
    tree.delete(*tree.get_children())
    total = 0
    for i, job in enumerate(filter_jobs()):
        tree.insert("", "end", iid=i, values=(
            job["date"],
            job["clinic"],
            job["doctor"],
            job["patient_name"],
            job["patient_surname"],
            job["prosthesis"],
            job["count"],
            f"₺{job['total_price']}",
            job["note"],
            job.get("status", "Hazırlanıyor")
        ))
        try:
            total += float(job["total_price"])
        except:
            pass
    total_label.config(text=f"Toplam Ciro: ₺{total:.2f}")


def show_qr_code():
    selected = tree.selection()
    if not selected:
        messagebox.showwarning("Uyarı", "QR kodu göstermek için bir kayıt seçin.")
        return

    idx = int(selected[0])
    job = filter_jobs()[idx]
    adsoyad = f"{job['patient_name']} {job['patient_surname']}"
    qr_data = (
        f"Hasta: {adsoyad}\n"
        f"Klinik: {job['clinic']}\n"
        f"Doktor: {job['doctor']}\n"
        f"Protez: {job['prosthesis']}\n"
        f"Üye: {job['count']}\n"
        f"Tarih: {job['date']}\n"
        f"Toplam: ₺{job['total_price']}\n"
        f"Not: {job['note']}"
    )

    filename = f"qr_{adsoyad}_{job['date'].replace('/', '-')}.png".replace(" ", "_")
    qr_path = os.path.join("data/qrcodes", filename)
    qr = qrcode.make(qr_data)
    qr.save(qr_path)

    # Görseli ve metni PIL ile birleştir
    qr_img = Image.open(qr_path).resize((300, 300))
    metin = qr_data.split('\n')
    font_path = "arial.ttf"  # Windows için
    try:
        from PIL import ImageDraw, ImageFont
        font = ImageFont.truetype(font_path, 18)
    except:
        font = None

    line_height = 30
    text_height = line_height * len(metin)
    final_img = Image.new("RGB", (qr_img.width, qr_img.height + text_height + 20), "white")
    final_img.paste(qr_img, (0, 0))

    draw = ImageDraw.Draw(final_img)
    for i, line in enumerate(metin):
        draw.text((10, qr_img.height + 10 + i * line_height), line, fill="black", font=font)

    final_path = os.path.join("data/qrcodes", f"printable_{filename}")
    final_img.save(final_path)

    # Ekranda göster
    qr_win = tk.Toplevel()
    qr_win.title("QR ve Bilgiler")

    img_tk = ImageTk.PhotoImage(final_img)
    panel = tk.Label(qr_win, image=img_tk)
    panel.image = img_tk
    panel.pack(padx=10, pady=10)

    # Yazdır
    def yazdir():
        try:
            os.startfile(final_path, "print")
        except Exception as e:
            messagebox.showerror("Hata", f"Yazdırma sırasında hata oluştu:\n{e}")

    tk.Button(qr_win, text="Yazdır", command=yazdir, bg="black", fg="white").pack(pady=10)

def delete_job():
    selected = tree.selection()
    if selected:
        idx = int(selected[0])
        job = filter_jobs()[idx]
        jobs.remove(job)
        save_data(FILES["jobs"], jobs)
        refresh_jobs()

def edit_job():
    global editing_index
    selected = tree.selection()
    if selected:
        idx = int(selected[0])
        job = filter_jobs()[idx]
        editing_index = jobs.index(job)
        name_var.set(job["patient_name"])
        surname_var.set(job["patient_surname"])
        clinic_var.set(job["clinic"])
        update_doctor_list()
        doctor_var.set(job["doctor"])
        prosthesis_var.set(job["prosthesis"])
        member_count_var.set(job["count"])
        note_var.set(job["note"])
        date_var.set(job["date"])
        total_price_var.set(job["total_price"])
        notebook.select(new_job_tab)
        job_status_var.set(job.get("status", "Hazırlanıyor"))


def export_excel():
    filtered = filter_jobs()
    if not filtered:
        messagebox.showwarning("Uyarı", "Dışa aktarılacak veri yok.")
        return
    path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Dosyası", "*.xlsx")])
    if path:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Tarih", "Klinik", "Doktor", "Ad", "Soyad", "Protez", "Üye", "Fiyat", "Not"])
        for job in filtered:
            ws.append([
                job["date"], job["clinic"], job["doctor"], job["patient_name"], job["patient_surname"],
                job["prosthesis"], job["count"], job["total_price"], job["note"]
            ])
        wb.save(path)
        messagebox.showinfo("Başarılı", f"Excel dosyası oluşturuldu:\n{path}")

def export_pdf():
    filtered = filter_jobs()
    if not filtered:
        messagebox.showwarning("Uyarı", "Dışa aktarılacak veri yok.")
        return

    path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Dosyası", "*.pdf")])
    if not path:
        return

    # Türkçe karakterleri düzgün göstermek için özel font kullan
    pdf = FPDF()
    pdf.add_page()

    # Arial yerine UTF-8 uyumlu bir font kullanmak gerekir (örnek: DejaVuSans)
    font_path = os.path.join("data", "DejaVuSans.ttf")
    if not os.path.exists(font_path):
        messagebox.showerror("Hata", f"PDF için gerekli font bulunamadı:\n{font_path}")
        return

    pdf.add_font("DejaVu", "", font_path, uni=True)
    pdf.set_font("DejaVu", size=10)
    pdf.cell(200, 10, "DP Lab - İş Listesi", ln=True, align="C")
    pdf.ln(5)

    for job in filtered:
        line = f"{job['date']} | {job['clinic']} | {job['doctor']} | {job['patient_name']} {job['patient_surname']} | {job['prosthesis']} | {job['count']} | ₺{job['total_price']} | {job['note']}"
        pdf.multi_cell(0, 8, line)
        pdf.ln(1)

    pdf.output(path)
    messagebox.showinfo("Başarılı", f"PDF dosyası oluşturuldu:\n{path}")


filters = tk.Frame(all_jobs_tab)
filters.pack(pady=5)
tk.Label(filters, text="Klinik").grid(row=0, column=0)
tk.Entry(filters, textvariable=filter_clinic, width=15).grid(row=0, column=1)
tk.Label(filters, text="Doktor").grid(row=0, column=2)
tk.Entry(filters, textvariable=filter_doctor, width=15).grid(row=0, column=3)
tk.Label(filters, text="Ad/Soyad").grid(row=0, column=11)
tk.Entry(filters, textvariable=filter_name, width=15).grid(row=0, column=12)

tk.Label(filters, text="Başlangıç Tarihi").grid(row=0, column=4)
tk.Entry(filters, textvariable=filter_start, width=12).grid(row=0, column=5)
tk.Label(filters, text="Bitiş Tarihi").grid(row=0, column=6)
tk.Entry(filters, textvariable=filter_end, width=12).grid(row=0, column=7)
tk.Button(filters, text="Filtrele", command=refresh_jobs).grid(row=0, column=8, padx=5)
tk.Button(filters, text="PDF Aktar", command=export_pdf).grid(row=0, column=9)
tk.Button(filters, text="Excel Aktar", command=export_excel).grid(row=0, column=10)

columns = ("Tarih", "Klinik", "Doktor", "Ad", "Soyad", "Protez", "Üye", "Fiyat", "Not", "Durum")
tree = ttk.Treeview(all_jobs_tab, columns=columns, show="headings")
for col in columns:
    tree.heading(col, text=col)
    tree.column(col, width=100)
tree.pack(fill="both", expand=True, padx=10, pady=5)

tree.heading("Durum", text="İş Durumu")
tree.column("Durum", width=120)

btns = tk.Frame(all_jobs_tab)
btns.pack()
tk.Button(btns, text="Seçiliyi Sil", command=delete_job, fg="red").pack(side="left", padx=10)
tk.Button(btns, text="Düzenle", command=edit_job).pack(side="left")

total_label = tk.Label(all_jobs_tab, text="Toplam Ciro: ₺0.00", font=("Arial", 12))
total_label.pack(pady=5)

refresh_jobs()
# === KLİNİK & DOKTOR SEKME ===

def refresh_clinic_list():
    clinic_listbox.delete(0, tk.END)
    for c in clinics:
        clinic_listbox.insert(tk.END, c["name"])

def refresh_doctor_list():
    doctor_listbox.delete(0, tk.END)
    for d in doctors:
        doctor_listbox.insert(tk.END, f'{d["name"]} ({d["clinic"]})')

def add_clinic():
    name = clinic_entry.get().strip()
    if name:
        clinics.append({"name": name})
        save_data(FILES["clinics"], clinics)
        refresh_clinic_list()
        clinic_entry.delete(0, tk.END)
        clinic_combo['values'] = [c["name"] for c in clinics]
        doctor_clinic_combo['values'] = [c["name"] for c in clinics]
        finance_clinic_combo['values'] = [c["name"] for c in clinics]  # <<< BU SATIR
        price_clinic_combo['values'] = ["Genel"] + [c["name"] for c in clinics]


        

def delete_clinic():
    selected = clinic_listbox.curselection()
    if selected:
        name = clinic_listbox.get(selected[0])
        clinics[:] = [c for c in clinics if c["name"] != name]
        save_data(FILES["clinics"], clinics)
        refresh_clinic_list()
        clinic_combo['values'] = [c["name"] for c in clinics]
        doctor_clinic_combo['values'] = [c["name"] for c in clinics]

def add_doctor():
    name = doctor_entry.get().strip()
    clinic = doctor_clinic_var.get()
    if name and clinic:
        doctors.append({"name": name, "clinic": clinic})
        save_data(FILES["doctors"], doctors)
        refresh_doctor_list()
        doctor_entry.delete(0, tk.END)
        doctor_combo['values'] = [d["name"] for d in doctors]

def delete_doctor():
    selected = doctor_listbox.curselection()
    if selected:
        full = doctor_listbox.get(selected)
        name = full.split(" (")[0]
        doctors[:] = [d for d in doctors if d["name"] != name]
        save_data(FILES["doctors"], doctors)
        refresh_doctor_list()
        doctor_combo['values'] = [d["name"] for d in doctors]

tk.Label(clinic_tab, text="Klinik Adı:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
clinic_entry = tk.Entry(clinic_tab)
clinic_entry.grid(row=0, column=1)
tk.Button(clinic_tab, text="Ekle", command=add_clinic).grid(row=0, column=2)
tk.Button(clinic_tab, text="Sil", command=delete_clinic).grid(row=0, column=3)

clinic_listbox = tk.Listbox(clinic_tab, height=6)
clinic_listbox.grid(row=1, column=0, columnspan=4, sticky="we", padx=10, pady=5)
refresh_clinic_list()

tk.Label(clinic_tab, text="Doktor Adı:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
doctor_entry = tk.Entry(clinic_tab)
doctor_entry.grid(row=2, column=1)

tk.Label(clinic_tab, text="Klinik Seç:").grid(row=2, column=2)
doctor_clinic_var = tk.StringVar()
doctor_clinic_combo = ttk.Combobox(clinic_tab, textvariable=doctor_clinic_var, values=[c["name"] for c in clinics])
doctor_clinic_combo.grid(row=2, column=3)

tk.Button(clinic_tab, text="Ekle", command=add_doctor).grid(row=3, column=1, pady=5)
tk.Button(clinic_tab, text="Sil", command=delete_doctor).grid(row=3, column=2, pady=5)

doctor_listbox = tk.Listbox(clinic_tab, height=6)
doctor_listbox.grid(row=4, column=0, columnspan=4, sticky="we", padx=10, pady=5)
refresh_doctor_list()

# === FİYAT LİSTESİ SEKME ===

def refresh_price_list():
    price_tree.delete(*price_tree.get_children())
    for p in prices:
        price_tree.insert("", "end", values=(p["type"], f"₺{p['price']}", p["clinic"]))


def add_price():
    tip = price_type_var.get().strip()
    fiyat = price_value_var.get().strip()
    klinik = price_clinic_var.get().strip()
    if tip and fiyat:
        prices.append({"type": tip, "price": float(fiyat), "clinic": klinik})
        save_data(FILES["prices"], prices)
        refresh_price_list()
        prosthesis_combo['values'] = list(set([p["type"] for p in prices]))


def delete_price():
    selected = price_tree.selection()
    if selected:
        item = price_tree.item(selected[0])["values"]
        selected_type = item[0]
        selected_clinic = item[2]
        prices[:] = [p for p in prices if not (p["type"] == selected_type and p["clinic"] == selected_clinic)]
        save_data(FILES["prices"], prices)
        refresh_price_list()
        prosthesis_combo['values'] = list(set([p["type"] for p in prices]))


def edit_price():
    selected = price_tree.selection()
    if selected:
        item = price_tree.item(selected[0])["values"]
        selected_type = item[0]
        selected_clinic = item[2]
        for p in prices:
            if p["type"] == selected_type and p["clinic"] == selected_clinic:
                try:
                    p["price"] = float(price_value_var.get())
                    save_data(FILES["prices"], prices)
                    refresh_price_list()
                    prosthesis_combo['values'] = list(set([p["type"] for p in prices]))
                except:
                    messagebox.showerror("Hata", "Fiyat sayısal olmalı.")
                break


tk.Label(price_tab, text="Protez Tipi:").grid(row=0, column=0, padx=10, pady=5)
price_type_var = tk.StringVar()
tk.Entry(price_tab, textvariable=price_type_var).grid(row=0, column=1)

tk.Label(price_tab, text="Fiyat (₺):").grid(row=0, column=2)
price_value_var = tk.StringVar()
tk.Entry(price_tab, textvariable=price_value_var).grid(row=0, column=3)

tk.Label(price_tab, text="Klinik:").grid(row=0, column=4)
price_clinic_var = tk.StringVar(value="Genel")
price_clinic_combo = ttk.Combobox(price_tab, textvariable=price_clinic_var, values=["Genel"] + [c["name"] for c in clinics])
price_clinic_combo.grid(row=0, column=5)


# Ekle / Sil / Düzenle butonlarını yatay yerleştirme
price_button_frame = tk.Frame(price_tab)
price_button_frame.grid(row=0, column=6, columnspan=2, padx=5)

tk.Button(price_button_frame, text="Ekle", command=add_price, bg="green", fg="white").pack(side="left", padx=3)
tk.Button(price_button_frame, text="Sil", command=delete_price, bg="red", fg="white").pack(side="left", padx=3)
tk.Button(price_button_frame, text="Düzenle", command=edit_price, bg="blue", fg="white").pack(side="left", padx=3)


price_tree = ttk.Treeview(price_tab, columns=("Tip", "Fiyat", "Klinik"), show="headings")
price_tree.heading("Tip", text="Protez Tipi")
price_tree.heading("Fiyat", text="Fiyat (₺)")
price_tree.heading("Klinik", text="Klinik")
price_tree.column("Tip", width=150)
price_tree.column("Fiyat", width=100)
price_tree.column("Klinik", width=150)
price_tree.grid(row=1, column=0, columnspan=7, padx=10, pady=10, sticky="nsew")


refresh_price_list()

tk.Button(btns, text="QR Kod Göster", command=show_qr_code).pack(side="left", padx=10)


# === MUHASEBE SEKME ===

finance_records = load_data("data/finance.json")
editing_finance_index = None

def save_finance():
    global editing_finance_index
    record = {
        "clinic": finance_clinic_var.get(),
        "type": finance_type_var.get(),
        "desc": finance_desc_var.get(),
        "amount": float(finance_amount_var.get()),
        "date": finance_date_var.get()
    }

    if editing_finance_index is None:
        finance_records.append(record)
    else:
        finance_records[editing_finance_index] = record
        editing_finance_index = None
        save_button.config(text="Kaydet", command=save_finance)

    save_data("data/finance.json", finance_records)
    refresh_finance()
    clear_finance_form()

def clear_finance_form():
    finance_clinic_var.set("")
    finance_type_var.set("Gelir")
    finance_desc_var.set("")
    finance_amount_var.set("")
    finance_date_var.set(datetime.today().strftime("%d/%m/%Y"))

def filtered_finance():
    result = []
    for r in finance_records:
        if finance_filter_clinic.get() and finance_filter_clinic.get() != r["clinic"]:
            continue
        if finance_filter_start.get():
            if datetime.strptime(r["date"], "%d/%m/%Y") < datetime.strptime(finance_filter_start.get(), "%d/%m/%Y"):
                continue
        if finance_filter_end.get():
            if datetime.strptime(r["date"], "%d/%m/%Y") > datetime.strptime(finance_filter_end.get(), "%d/%m/%Y"):
                continue
        result.append(r)
    return result

def refresh_finance():
    finance_tree.delete(*finance_tree.get_children())
    total_income = 0
    total_expense = 0
    for idx, r in enumerate(filtered_finance()):
        finance_tree.insert("", "end", iid=idx, values=(r["date"], r["clinic"], r["type"], r["desc"], f"₺{r['amount']}"))
        if r["type"] == "Gelir":
            total_income += r["amount"]
        else:
            total_expense += r["amount"]
    net = total_income - total_expense
    finance_total_label.config(text=f"Toplam Gelir: ₺{total_income:.2f} | Gider: ₺{total_expense:.2f} | Kalan: ₺{net:.2f}")

    

def delete_finance():
    selected = finance_tree.selection()
    if selected:
        index = int(selected[0])
        record = filtered_finance()[index]
        for i, r in enumerate(finance_records):
            if r == record:
                del finance_records[i]
                break
        save_data("data/finance.json", finance_records)
        refresh_finance()

def edit_finance():
    global editing_finance_index
    selected = finance_tree.selection()
    if selected:
        index = int(selected[0])
        record = filtered_finance()[index]
        for i, r in enumerate(finance_records):
            if r == record:
                editing_finance_index = i
                break

        finance_clinic_var.set(record["clinic"])
        finance_type_var.set(record["type"])
        finance_desc_var.set(record["desc"])
        finance_amount_var.set(str(record["amount"]))
        finance_date_var.set(record["date"])
        save_button.config(text="Güncelle", command=save_finance)

        

# === GİRİŞ ALANLARI ===
finance_clinic_var = tk.StringVar()
finance_type_var = tk.StringVar(value="Gelir")
finance_desc_var = tk.StringVar()
finance_amount_var = tk.StringVar()
finance_date_var = tk.StringVar(value=datetime.today().strftime("%d/%m/%Y"))

tk.Label(finance_tab, text="Klinik:").grid(row=0, column=0)
finance_clinic_combo = ttk.Combobox(finance_tab, textvariable=finance_clinic_var, values=[c["name"] for c in clinics])
finance_clinic_combo.grid(row=0, column=1)

tk.Label(finance_tab, text="Tür:").grid(row=0, column=2)
finance_type_combo = ttk.Combobox(finance_tab, textvariable=finance_type_var, values=["Gelir", "Gider"])
finance_type_combo.grid(row=0, column=3)

tk.Label(finance_tab, text="Açıklama:").grid(row=1, column=0)
tk.Entry(finance_tab, textvariable=finance_desc_var).grid(row=1, column=1, columnspan=3, sticky="we", padx=5)

tk.Label(finance_tab, text="Tutar (₺):").grid(row=2, column=0)
tk.Entry(finance_tab, textvariable=finance_amount_var).grid(row=2, column=1)

tk.Label(finance_tab, text="Tarih:").grid(row=2, column=2)
tk.Entry(finance_tab, textvariable=finance_date_var).grid(row=2, column=3)

save_button = tk.Button(finance_tab, text="Kaydet", command=save_finance, bg="green", fg="white")
save_button.grid(row=3, column=3, pady=10)

# === FİLTRE ALANI ===
finance_filter_clinic = tk.StringVar()
finance_filter_start = tk.StringVar()
finance_filter_end = tk.StringVar()

tk.Label(finance_tab, text="Filtre Klinik:").grid(row=4, column=0)
tk.Entry(finance_tab, textvariable=finance_filter_clinic).grid(row=4, column=1)
tk.Label(finance_tab, text="Başlangıç Tarihi:").grid(row=4, column=2)
tk.Entry(finance_tab, textvariable=finance_filter_start).grid(row=4, column=3)
tk.Label(finance_tab, text="Bitiş Tarihi:").grid(row=5, column=2)
tk.Entry(finance_tab, textvariable=finance_filter_end).grid(row=5, column=3)
tk.Button(finance_tab, text="Filtrele", command=refresh_finance).grid(row=5, column=1, pady=5)

# === KAYIT LİSTESİ ===
finance_tree = ttk.Treeview(finance_tab, columns=("Tarih", "Klinik", "Tür", "Açıklama", "Tutar"), show="headings")
for col in ("Tarih", "Klinik", "Tür", "Açıklama", "Tutar"):
    finance_tree.heading(col, text=col)
    finance_tree.column(col, width=150)
finance_tree.grid(row=6, column=0, columnspan=4, padx=10, pady=10)

# === BUTONLAR ===
btn_frame = tk.Frame(finance_tab)
btn_frame.grid(row=7, column=0, columnspan=4, pady=10)

tk.Button(btn_frame, text="Seçiliyi Sil", command=delete_finance, fg="red").pack(side="left", padx=10)
tk.Button(btn_frame, text="Düzenle", command=edit_finance).pack(side="left", padx=10)

finance_total_label = tk.Label(finance_tab, text="Toplam Gelir: ₺0.00 | Gider: ₺0.00 | Kalan: ₺0.00", font=("Arial", 11, "bold"))
finance_total_label.grid(row=8, column=0, columnspan=4, pady=10)

refresh_finance()

# === BORÇ TAKİBİ SEKME ===

borc_tab = ttk.Frame(notebook)
notebook.add(borc_tab, text="Borç Takibi")

def hesapla_borclar():
    borclar = []
    for clinic in clinics:
        ciro = sum(float(j["total_price"]) for j in jobs if j["clinic"] == clinic["name"])
        odenen = sum(float(r["amount"]) for r in finance_records if r["clinic"] == clinic["name"] and r["type"] == "Gelir")
        kalan = ciro - odenen
        borclar.append({
            "clinic": clinic["name"],
            "ciro": ciro,
            "odeme": odenen,
            "borc": kalan
        })
    return borclar

def guncelle_borc_tablosu():
    borc_tree.delete(*borc_tree.get_children())
    for borc in hesapla_borclar():
        borc_tree.insert("", "end", values=(
            borc["clinic"],
            f"₺{borc['ciro']:.2f}",
            f"₺{borc['odeme']:.2f}",
            f"₺{borc['borc']:.2f}"
        ))

borc_tree = ttk.Treeview(borc_tab, columns=("Klinik", "Ciro", "Ödenen", "Kalan Borç"), show="headings")
for col in ("Klinik", "Ciro", "Ödenen", "Kalan Borç"):
    borc_tree.heading(col, text=col)
    borc_tree.column(col, width=200)
borc_tree.pack(fill="both", expand=True, padx=10, pady=10)

tk.Button(borc_tab, text="Borçları Güncelle", command=guncelle_borc_tablosu).pack(pady=10)

guncelle_borc_tablosu()

def hesapla_klinik_ciro():
    ciro_dict = {}
    for job in jobs:
        klinik = job["clinic"]
        fiyat = float(job.get("total_price", 0))
        ciro_dict[klinik] = ciro_dict.get(klinik, 0) + fiyat
    return ciro_dict

def hesapla_aylik_ciro():
    aylik_dict = {}
    for job in jobs:
        try:
            tarih = datetime.strptime(job["date"], "%d/%m/%Y")
            ay = tarih.strftime("%Y-%m")
            fiyat = float(job.get("total_price", 0))
            aylik_dict[ay] = aylik_dict.get(ay, 0) + fiyat
        except:
            continue
    return aylik_dict

def guncelle_raporlar():
    clinic_tree.delete(*clinic_tree.get_children())
    month_tree.delete(*month_tree.get_children())

    klinik_ciro = hesapla_klinik_ciro()
    aylik_ciro = hesapla_aylik_ciro()

    for k, t in klinik_ciro.items():
        clinic_tree.insert("", "end", values=(k, f"₺{t:.2f}"))

    for a, t in sorted(aylik_ciro.items()):
        month_tree.insert("", "end", values=(a, f"₺{t:.2f}"))


# === LABORATUVAR ENVANTERİ SEKME ===

envanter_tab = ttk.Frame(notebook)
notebook.add(envanter_tab, text="Lab Envanteri")

envanter_kayitlari = load_data("data/envanter.json")
editing_envanter_index = None

# Değişkenler
env_ad_var = tk.StringVar()
env_miktar_var = tk.StringVar()
env_birim_var = tk.StringVar()
env_giris_var = tk.StringVar(value=datetime.today().strftime("%d/%m/%Y"))
env_skt_var = tk.StringVar()
env_siparis_var = tk.StringVar()
env_not_var = tk.StringVar()

# Kaydet/Güncelle
def kaydet_envanter():
    global editing_envanter_index
    try:
        kayit = {
            "ad": env_ad_var.get(),
            "miktar": float(env_miktar_var.get()),
            "birim": env_birim_var.get(),
            "giris": env_giris_var.get(),
            "skt": env_skt_var.get(),
            "siparis": env_siparis_var.get(),
            "not": env_not_var.get()
        }

        if editing_envanter_index is None:
            envanter_kayitlari.append(kayit)
        else:
            envanter_kayitlari[editing_envanter_index] = kayit
            editing_envanter_index = None
            kaydet_button.config(text="Ekle", command=kaydet_envanter)

        save_data("data/envanter.json", envanter_kayitlari)
        guncelle_envanter()
        temizle_envanter()
    except:
        messagebox.showerror("Hata", "Tüm alanları doğru şekilde doldurunuz.")

def temizle_envanter():
    global editing_envanter_index
    env_ad_var.set("")
    env_miktar_var.set("")
    env_birim_var.set("")
    env_giris_var.set(datetime.today().strftime("%d/%m/%Y"))
    env_skt_var.set("")
    env_siparis_var.set("")
    env_not_var.set("")
    editing_envanter_index = None
    kaydet_button.config(text="Ekle", command=kaydet_envanter)

# Listeyi yenile
def guncelle_envanter():
    envanter_tree.delete(*envanter_tree.get_children())
    for i, k in enumerate(envanter_kayitlari):
        envanter_tree.insert("", "end", iid=i, values=(
            k["ad"], k["miktar"], k["birim"], k["giris"],
            k["skt"], k["siparis"], k["not"]
        ))

# Sil
def sil_envanter():
    selected = envanter_tree.selection()
    if selected:
        idx = int(selected[0])
        del envanter_kayitlari[idx]
        save_data("data/envanter.json", envanter_kayitlari)
        guncelle_envanter()
        temizle_envanter()

# Düzenle
def duzenle_envanter():
    global editing_envanter_index
    selected = envanter_tree.selection()
    if selected:
        idx = int(selected[0])
        kayit = envanter_kayitlari[idx]
        editing_envanter_index = idx

        env_ad_var.set(kayit["ad"])
        env_miktar_var.set(str(kayit["miktar"]))
        env_birim_var.set(kayit["birim"])
        env_giris_var.set(kayit["giris"])
        env_skt_var.set(kayit["skt"])
        env_siparis_var.set(kayit["siparis"])
        env_not_var.set(kayit["not"])

        kaydet_button.config(text="Güncelle", command=kaydet_envanter)

# GİRİŞ ALANI
tk.Label(envanter_tab, text="Ad").grid(row=0, column=0)
tk.Entry(envanter_tab, textvariable=env_ad_var).grid(row=0, column=1)

tk.Label(envanter_tab, text="Miktar").grid(row=0, column=2)
tk.Entry(envanter_tab, textvariable=env_miktar_var).grid(row=0, column=3)

tk.Label(envanter_tab, text="Birim").grid(row=1, column=0)
tk.Entry(envanter_tab, textvariable=env_birim_var).grid(row=1, column=1)

tk.Label(envanter_tab, text="Giriş Tarihi").grid(row=1, column=2)
tk.Entry(envanter_tab, textvariable=env_giris_var).grid(row=1, column=3)

tk.Label(envanter_tab, text="SKT").grid(row=2, column=0)
tk.Entry(envanter_tab, textvariable=env_skt_var).grid(row=2, column=1)

tk.Label(envanter_tab, text="Sipariş Tarihi").grid(row=2, column=2)
tk.Entry(envanter_tab, textvariable=env_siparis_var).grid(row=2, column=3)

tk.Label(envanter_tab, text="Not").grid(row=3, column=0)
tk.Entry(envanter_tab, textvariable=env_not_var, width=40).grid(row=3, column=1, columnspan=3)

# BUTONLAR
kaydet_button = tk.Button(envanter_tab, text="Ekle", command=kaydet_envanter, bg="green", fg="white")
kaydet_button.grid(row=4, column=1, pady=5)

tk.Button(envanter_tab, text="Düzenle", command=duzenle_envanter).grid(row=4, column=2)
tk.Button(envanter_tab, text="Seçiliyi Sil", command=sil_envanter, fg="red").grid(row=4, column=3)

# LİSTE
envanter_tree = ttk.Treeview(envanter_tab, columns=("Ad", "Miktar", "Birim", "Giriş", "SKT", "Sipariş", "Not"), show="headings")
for col in ("Ad", "Miktar", "Birim", "Giriş", "SKT", "Sipariş", "Not"):
    envanter_tree.heading(col, text=col)
    envanter_tree.column(col, width=130)
envanter_tree.grid(row=5, column=0, columnspan=4, padx=10, pady=10)

guncelle_envanter()


tk.Label(report_tab, text="Klinik Bazlı Ciro", font=("Arial", 12, "bold")).pack(pady=5)
clinic_tree = ttk.Treeview(report_tab, columns=("Klinik", "Ciro"), show="headings")
clinic_tree.heading("Klinik", text="Klinik")
clinic_tree.heading("Ciro", text="Toplam Ciro (₺)")
clinic_tree.column("Klinik", width=200)
clinic_tree.column("Ciro", width=150)
clinic_tree.pack(pady=5)

tk.Label(report_tab, text="Aylık Gelir", font=("Arial", 12, "bold")).pack(pady=5)
month_tree = ttk.Treeview(report_tab, columns=("Ay", "Toplam"), show="headings")
month_tree.heading("Ay", text="Ay")
month_tree.heading("Toplam", text="Toplam Gelir (₺)")
month_tree.column("Ay", width=200)
month_tree.column("Toplam", width=150)
month_tree.pack(pady=5)
def export_report_pdf():
    path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Dosyası", "*.pdf")])
    if not path:
        return

    font_path = os.path.join("data", "DejaVuSans.ttf")
    if not os.path.exists(font_path):
        messagebox.showerror("Hata", f"PDF için gerekli font bulunamadı:\n{font_path}")
        return

    pdf = FPDF()
    pdf.add_page()
    pdf.add_font("DejaVu", "", font_path, uni=True)
    pdf.set_font("DejaVu", size=10)

    pdf.cell(0, 10, "Klinik Bazlı Ciro", ln=True)
    for k, v in hesapla_klinik_ciro().items():
        pdf.cell(0, 8, f"{k}: ₺{v:.2f}", ln=True)

    pdf.ln(5)
    pdf.cell(0, 10, "Aylık Ciro", ln=True)
    for a, v in sorted(hesapla_aylik_ciro().items()):
        pdf.cell(0, 8, f"{a}: ₺{v:.2f}", ln=True)

    pdf.output(path)
    messagebox.showinfo("Başarılı", f"PDF raporu kaydedildi:\n{path}")


def export_report_excel():
    path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Dosyası", "*.xlsx")])
    if not path:
        return

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Klinik Ciro"
    ws1.append(["Klinik", "Toplam Ciro (₺)"])
    for k, v in hesapla_klinik_ciro().items():
        ws1.append([k, v])

    ws2 = wb.create_sheet("Aylık Ciro")
    ws2.append(["Ay", "Toplam Gelir (₺)"])
    for a, v in sorted(hesapla_aylik_ciro().items()):
        ws2.append([a, v])

    wb.save(path)
    messagebox.showinfo("Başarılı", f"Excel raporu kaydedildi:\n{path}")

btn_frame = tk.Frame(report_tab)
btn_frame.pack(pady=10)

tk.Button(btn_frame, text="PDF Aktar", command=export_report_pdf, bg="black", fg="white").pack(side="left", padx=10)
tk.Button(btn_frame, text="Excel Aktar", command=export_report_excel, bg="green", fg="white").pack(side="left", padx=10)


guncelle_raporlar()
def export_report_excel():
    path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Dosyası", "*.xlsx")])
    if not path:
        return

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Klinik Ciro"
    ws1.append(["Klinik", "Toplam Ciro (₺)"])
    for k, v in hesapla_klinik_ciro().items():
        ws1.append([k, v])

    ws2 = wb.create_sheet("Aylık Ciro")
    ws2.append(["Ay", "Toplam Gelir (₺)"])
    for a, v in sorted(hesapla_aylik_ciro().items()):
        ws2.append([a, v])

    wb.save(path)
    messagebox.showinfo("Başarılı", f"Excel raporu kaydedildi:\n{path}")

def export_report_pdf():
    path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Dosyası", "*.pdf")])
    if not path:
        return

    font_path = os.path.join("data", "DejaVuSans.ttf")
    if not os.path.exists(font_path):
        messagebox.showerror("Hata", f"PDF için gerekli font bulunamadı:\n{font_path}")
        return

    pdf = FPDF()
    pdf.add_page()
    pdf.add_font("DejaVu", "", font_path, uni=True)
    pdf.set_font("DejaVu", size=10)

    pdf.cell(0, 10, "Klinik Bazlı Ciro", ln=True)
    for k, v in hesapla_klinik_ciro().items():
        pdf.cell(0, 8, f"{k}: ₺{v:.2f}", ln=True)

    pdf.ln(5)
    pdf.cell(0, 10, "Aylık Ciro", ln=True)
    for a, v in sorted(hesapla_aylik_ciro().items()):
        pdf.cell(0, 8, f"{a}: ₺{v:.2f}", ln=True)

    pdf.output(path)
    messagebox.showinfo("Başarılı", f"PDF raporu kaydedildi:\n{path}")


# === UYGULAMA BAŞLAT ===
root.mainloop()
